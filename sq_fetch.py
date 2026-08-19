#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
日経225オプション SQレポート データ取得層

JPXが公表する日次相場情報から、行使価格ごとの建玉・清算値段・出来高を取り出す。

設計方針:
  * JPXのファイル名・配置ディレクトリはCMSの都合で変わるため、URLを決め打ちしない。
    一覧ページのHTMLから `ose20260818.zip` のようなリンクを拾って最新を選ぶ。
  * CSVの列順も決め打ちしない。ヘッダ行を探して列名（日本語/英語）から列位置を判定する。
    列が増減しても壊れないようにする。
  * 取得できなかった項目は例外にせず None を返し、レポート側で
    「データ品質の開示」として明示する。元レポートと同じ運用にする。

データの所在（2026-08 に実地調査した結果）:
  * 行使価格別の建玉は、日次では大阪取引所日報のPDF（siop_dyr_*.pdf）にしか無い。
    建玉残高のExcelは日次公開ではなく、日次相場情報のページにはファイルが置かれていない。
  * 清算値段・IV・原資産価格・金利・残日数は rbYYYYMMDD.csv で取れる。
  * 日報は当日中に公開されないことがあるため、公開済みの最新営業日を採用する。

取得元を差し替えたい場合:
  SQ_CHAIN_FILE=path/to/225OP.xlsx   手元のオプション板ファイルを直接読む
  SQ_NIKKEI_VI=33.6                  日経VI（自動取得できない場合の手入力）
"""

from __future__ import annotations

import csv
import io
import os
import re
import sys
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import sq_analytics as sa

USER_AGENT = ("Mozilla/5.0 (compatible; market-dashboard/1.0; "
              "+https://github.com/komineriko/market-dashboard)")
JPX_DAILY_PAGE = "https://www.jpx.co.jp/markets/statistics-derivatives/daily/index.html"
# 清算値段等（先物・オプション）。行使価格ごとの清算価格・IV・原資産価格・金利・残日数を持つ
JPX_SETTLEMENT_PAGE = "https://www.jpx.co.jp/markets/derivatives/settlement-price/index.html"
JPX_BASE = "https://www.jpx.co.jp"
TIMEOUT = 60

# 日経225オプションを見分けるための銘柄・商品コードの手掛かり
NIKKEI_OPTION_HINTS = ("日経225オプション", "日経225オプシヨン", "NK225E", "NK225 OP",
                       "NIKKEI 225 OPTIONS", "225オプション")
NIKKEI_OPTION_EXCLUDE = ("ミニ", "MINI", "WEEKLY", "ウィークリー", "週次")


class FetchError(RuntimeError):
    """取得・解析に失敗し、レポートを組めない場合。"""


# ---------------------------------------------------------------------------
# HTTP
# ---------------------------------------------------------------------------

def _requests():
    import requests  # 遅延importにして、解析だけのテストでは不要にする
    return requests


def http_get(url: str, retries: int = 3) -> bytes:
    requests = _requests()
    last = None
    for attempt in range(retries):
        try:
            r = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=TIMEOUT)
            r.raise_for_status()
            return r.content
        except Exception as exc:      # noqa: BLE001 - リトライして最後に投げ直す
            last = exc
            if attempt < retries - 1:
                import time
                time.sleep(2 ** attempt)
    raise FetchError(f"取得に失敗しました: {url} ({last})")


# ---------------------------------------------------------------------------
# JPXの日次ファイルを一覧ページから探す
# ---------------------------------------------------------------------------

FILE_LINK_RE = re.compile(
    r'href="([^"]*?/(?:ose|OSE)(\d{8})\.(zip|csv|xls|xlsx))"', re.IGNORECASE)


def discover_daily_files(page_html: str) -> List[Tuple[date, str]]:
    """一覧ページのHTMLから (日付, 絶対URL) を新しい順に返す。"""
    found: Dict[date, str] = {}
    for href, ymd, _ext in FILE_LINK_RE.findall(page_html):
        try:
            d = datetime.strptime(ymd, "%Y%m%d").date()
        except ValueError:
            continue
        url = href if href.startswith("http") else JPX_BASE + (
            href if href.startswith("/") else "/" + href)
        found.setdefault(d, url)
    return sorted(found.items(), key=lambda x: x[0], reverse=True)


def latest_daily_file_url(target: Optional[date] = None) -> Tuple[date, str]:
    override = os.environ.get("SQ_JPX_DAILY_URL")
    if override:
        m = re.search(r"(\d{8})", override)
        d = datetime.strptime(m.group(1), "%Y%m%d").date() if m else date.today()
        return d, override
    html = http_get(JPX_DAILY_PAGE).decode("utf-8", errors="replace")
    files = discover_daily_files(html)
    if not files:
        raise FetchError(
            "JPXの日次相場情報ページからデータファイルのリンクを見つけられませんでした。"
            "ページ構成が変わった可能性があります（SQ_JPX_DAILY_URL で直接指定できます）。")
    if target:
        for d, url in files:
            if d == target:
                return d, url
        raise FetchError(f"{target} のファイルが一覧に見つかりません（最新は {files[0][0]}）。")
    return files[0]


# ---------------------------------------------------------------------------
# 表形式ファイルの読み込み（列名で列位置を決める）
# ---------------------------------------------------------------------------

# UTF-8 を先に試す。cp932 を先にすると UTF-8 のファイルが化けたまま成功しうる
ENCODINGS = ("utf-8-sig", "utf-8", "cp932", "euc-jp")

# 論理名 → 列名に現れうるキーワード（順に優先）
COLUMN_HINTS: Dict[str, Sequence[str]] = {
    "product":  ("銘柄", "商品", "限月取引", "コード", "product", "contract"),
    "month":    ("限月", "expir", "contract month", "maturity"),
    "strike":   ("権利行使価格", "行使価格", "strike"),
    "kind":     ("プット", "コール", "put/call", "putcall", "call/put", "種類"),
    "settle":   ("清算値段", "決済価格", "settlement"),
    # 「価格」単体は「権利行使価格」に誤爆するので入れない
    "close":    ("終値", "現在値", "プレミアム", "last", "close"),
    "oi":       ("建玉", "取組高", "open interest", "openinterest"),
    "volume":   ("出来高", "売買高", "volume"),
}


def decode_text(raw: bytes) -> str:
    for enc in ENCODINGS:
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("cp932", errors="replace")


def rows_from_bytes(raw: bytes, name: str = "") -> List[List[str]]:
    """zip / csv / xls(x) を行のリストに落とす。"""
    lower = name.lower()
    if raw[:2] == b"PK" and (lower.endswith(".zip") or not lower.endswith((".xlsx", ".xlsm"))):
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            inner = [n for n in zf.namelist() if not n.endswith("/")]
            if not inner:
                raise FetchError("ZIPが空です。")
            # 一番大きいファイルを本体とみなす
            inner.sort(key=lambda n: zf.getinfo(n).file_size, reverse=True)
            return rows_from_bytes(zf.read(inner[0]), inner[0])
    if lower.endswith((".xlsx", ".xlsm", ".xls")) or raw[:2] == b"PK":
        return _rows_from_excel(raw)
    text = decode_text(raw)
    return [r for r in csv.reader(io.StringIO(text))]


def sheets_from_excel(raw: bytes) -> List[Tuple[str, List[List[str]]]]:
    """Excelをシート単位で返す。限月ごとにシートが分かれている板ファイルに対応する。"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise FetchError("Excelファイルを読むには openpyxl が必要です。") from exc
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    out: List[Tuple[str, List[List[str]]]] = []
    for ws in wb.worksheets:
        rows = [["" if c is None else str(c) for c in row]
                for row in ws.iter_rows(values_only=True)]
        if any(any(c for c in r) for r in rows):
            out.append((ws.title, rows))
    return out


def _rows_from_excel(raw: bytes) -> List[List[str]]:
    out: List[List[str]] = []
    for _name, rows in sheets_from_excel(raw):
        out.extend(rows)
    return out


def find_header(rows: Sequence[Sequence[str]], required=("strike", "oi")) -> Optional[Tuple[int, Dict[str, int]]]:
    """
    ヘッダ行を探し、論理名 → 列インデックスの対応を返す。
    必須列（行使価格・建玉）が揃った最初の行を採用する。
    """
    for i, row in enumerate(rows[:80]):
        cells = [str(c).strip().lower().replace(" ", "").replace("　", "") for c in row]
        if not any(cells):
            continue
        mapping: Dict[str, int] = {}
        for logical, hints in COLUMN_HINTS.items():
            for j, cell in enumerate(cells):
                if not cell or j in mapping.values():
                    continue
                if any(h.lower().replace(" ", "") in cell for h in hints):
                    mapping[logical] = j
                    break
        if all(k in mapping for k in required):
            return i, mapping
    return None


def _num(value: str) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip().replace(",", "").replace("¥", "")
    if s in ("", "-", "--", "―", "‐", "*", "N/A", "na", "nan", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _is_call(text: str) -> Optional[bool]:
    t = str(text).strip().upper()
    if not t:
        return None
    if t.startswith("C") or "コール" in t or "CALL" in t:
        return True
    if t.startswith("P") or "プット" in t or "PUT" in t:
        return False
    return None


def _normalise_month(text: str) -> Optional[str]:
    """"2026/09" "202609" "26年9月" などを "26-09" に寄せる。"""
    s = str(text)
    m = re.search(r"(20\d{2})\D{0,3}(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)) % 100:02d}-{int(m.group(2)):02d}"
    m = re.search(r"\b(\d{2})\D?(\d{2})\b", s)
    if m:
        return f"{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return None


def parse_option_chains(raw: bytes, name: str = "",
                        months: Optional[Sequence[str]] = None) -> Dict[str, sa.Chain]:
    """JPX日次相場情報（または同等の板ファイル）から限月ごとの Chain を組み立てる。"""
    return parse_rows_long(rows_from_bytes(raw, name), months)


def parse_rows_long(rows: Sequence[Sequence[str]],
                    months: Optional[Sequence[str]] = None) -> Dict[str, sa.Chain]:
    """
    1行=1オプションの長形式を読む。列順に依存せず、ヘッダの列名から位置を判定する。
    """
    header = find_header(rows)
    if not header:
        raise FetchError(
            "行使価格・建玉の列を持つヘッダ行を見つけられませんでした。"
            "ファイル形式が想定と違う可能性があります。")
    idx, col = header
    chains: Dict[str, sa.Chain] = {}
    skipped_products = 0

    for row in rows[idx + 1:]:
        if not any(str(c).strip() for c in row):
            continue

        def cell(logical: str) -> str:
            j = col.get(logical)
            return str(row[j]).strip() if j is not None and j < len(row) else ""

        product = cell("product")
        if product:
            blob = product.upper()
            if any(x.upper() in blob for x in NIKKEI_OPTION_EXCLUDE):
                skipped_products += 1
                continue
            if NIKKEI_OPTION_HINTS and not any(h.upper() in blob for h in NIKKEI_OPTION_HINTS):
                # 商品名が分かるのに日経225オプションでない行は捨てる
                if any(ch.isalpha() or ord(ch) > 127 for ch in product):
                    skipped_products += 1
                    continue

        strike = _num(cell("strike"))
        if not strike or strike <= 0:
            continue          # 先物行など、行使価格を持たない行

        kind = _is_call(cell("kind"))
        if kind is None:
            continue

        month = _normalise_month(cell("month"))
        if month is None:
            continue
        if months and month not in months:
            continue

        price = _num(cell("settle"))
        if price is None:
            price = _num(cell("close"))
        oi = _num(cell("oi")) or 0
        vol = _num(cell("volume")) or 0

        chain = chains.setdefault(month, sa.Chain(contract_month=month))
        row_obj = chain.rows.get(strike) or sa.StrikeRow(strike=strike)
        if kind:
            row_obj.call_price = price
            row_obj.call_oi = int(oi)
            row_obj.call_volume = int(vol)
        else:
            row_obj.put_price = price
            row_obj.put_oi = int(oi)
            row_obj.put_volume = int(vol)
        chain.add(row_obj)

    if not chains:
        raise FetchError(
            "日経225オプションの行を1件も抽出できませんでした"
            f"（{skipped_products}行を対象外として除外）。")
    return chains


# ---------------------------------------------------------------------------
# 現物・日経VI
# ---------------------------------------------------------------------------

@dataclass
class SpotQuote:
    close: Optional[float] = None
    prev_close: Optional[float] = None
    change: Optional[float] = None
    change_pct: Optional[float] = None
    open: Optional[float] = None
    high: Optional[float] = None
    low: Optional[float] = None
    as_of: Optional[str] = None
    source: Optional[str] = None

    @property
    def ok(self) -> bool:
        return self.close is not None


def fetch_nikkei_spot(as_of: Optional[date] = None) -> SpotQuote:
    """
    日経平均の四本値。yfinance を第一候補、Stooq をフォールバックにする。

    as_of を渡すとその営業日の値を返す。板データが1営業日遅れで公開されるため、
    指定しないと板と現物で日付がずれて前日比が食い違う。
    """
    q = _spot_from_yfinance(as_of)
    if q.ok:
        return q
    return _spot_from_stooq(as_of)


def _pick_index(dates: Sequence, as_of: Optional[date]) -> int:
    """as_of に一致する行を返す。無ければ最後の行。"""
    if as_of is None:
        return len(dates) - 1
    for i in range(len(dates) - 1, -1, -1):
        if dates[i] == as_of:
            return i
    return len(dates) - 1


def _spot_from_yfinance(as_of: Optional[date] = None) -> SpotQuote:
    try:
        import yfinance as yf
        hist = yf.Ticker("^N225").history(period="1mo", auto_adjust=False)
        if hist is None or hist.empty:
            return SpotQuote()
        days = [d.date() for d in hist.index]
        i = _pick_index(days, as_of)
        if i <= 0:
            return SpotQuote()
        last = hist.iloc[i]
        prev = hist.iloc[i - 1]
        close = float(last["Close"])
        prev_close = float(prev["Close"]) if prev is not None else None
        return SpotQuote(
            close=close, prev_close=prev_close,
            change=(close - prev_close) if prev_close else None,
            change_pct=((close / prev_close - 1) * 100) if prev_close else None,
            open=float(last["Open"]), high=float(last["High"]), low=float(last["Low"]),
            as_of=str(days[i]), source="yfinance ^N225")
    except Exception:      # noqa: BLE001 - 取得できなければフォールバックへ
        return SpotQuote()


def _spot_from_stooq(as_of: Optional[date] = None) -> SpotQuote:
    try:
        raw = http_get("https://stooq.com/q/d/l/?s=^nkx&i=d")
        rows = list(csv.DictReader(io.StringIO(raw.decode("utf-8", errors="replace"))))
        if len(rows) < 2:
            return SpotQuote()
        days = [datetime.strptime(r["Date"], "%Y-%m-%d").date() for r in rows]
        i = _pick_index(days, as_of)
        if i <= 0:
            return SpotQuote()
        last, prev = rows[i], rows[i - 1]
        close, prev_close = float(last["Close"]), float(prev["Close"])
        return SpotQuote(
            close=close, prev_close=prev_close, change=close - prev_close,
            change_pct=(close / prev_close - 1) * 100,
            open=float(last["Open"]), high=float(last["High"]), low=float(last["Low"]),
            as_of=last["Date"], source="Stooq ^NKX")
    except Exception:      # noqa: BLE001
        return SpotQuote()


def fetch_nikkei_vi() -> Tuple[Optional[float], Optional[str]]:
    """
    日経VI。手入力（SQ_NIKKEI_VI）を最優先する。

    元レポートも日経VIは「引け直前の水準」を人が読み取って与えている。
    終値は 15:46 前後に歪むため、自動取得値をそのまま判断に使うのは危険で、
    取れなかった場合は素直に未取得として開示する。
    """
    manual = os.environ.get("SQ_NIKKEI_VI")
    if manual:
        v = _num(manual)
        if v:
            return v, "手入力（SQ_NIKKEI_VI）"
    try:
        import yfinance as yf
        for symbol in ("^N225VI", "^NKVI"):
            hist = yf.Ticker(symbol).history(period="5d")
            if hist is not None and not hist.empty:
                return float(hist.iloc[-1]["Close"]), f"yfinance {symbol}（終値）"
    except Exception:      # noqa: BLE001
        pass
    return None, None


# ---------------------------------------------------------------------------
# 横形式（CALL列 | 行使価格 | PUT列）の板ファイル
# ---------------------------------------------------------------------------

MONTH_IN_TEXT_RE = re.compile(r"(20\d{2})\s*年?\s*(\d{1,2})\s*月限|(\d{2})-(\d{2})")


def sniff_month(rows: Sequence[Sequence[str]], limit: int = 40) -> Optional[str]:
    """シート冒頭のタイトル行などから限月を拾う。横形式は限月列を持たないことが多い。"""
    for row in rows[:limit]:
        for cell in row:
            m = MONTH_IN_TEXT_RE.search(str(cell))
            if m:
                if m.group(1):
                    return f"{int(m.group(1)) % 100:02d}-{int(m.group(2)):02d}"
                return f"{int(m.group(3)):02d}-{int(m.group(4)):02d}"
    return None


def _column_positions(cells: Sequence[str], hints: Sequence[str]) -> List[int]:
    out = []
    for j, cell in enumerate(cells):
        if cell and any(h.lower().replace(" ", "") in cell for h in hints):
            out.append(j)
    return out


def parse_wide_chain(rows: Sequence[Sequence[str]],
                     default_month: Optional[str] = None) -> Optional[sa.Chain]:
    """
    行使価格を中央に置き、左をCALL・右をPUTとする板レイアウトを読む。
    建玉列が行使価格の左右に1つずつある場合にこの形式と判定する。
    """
    for i, row in enumerate(rows[:80]):
        cells = [str(c).strip().lower().replace(" ", "").replace("　", "") for c in row]
        if not any(cells):
            continue
        strikes = _column_positions(cells, COLUMN_HINTS["strike"])
        if len(strikes) != 1:
            continue
        s_idx = strikes[0]
        oi_cols = _column_positions(cells, COLUMN_HINTS["oi"])
        left_oi = [j for j in oi_cols if j < s_idx]
        right_oi = [j for j in oi_cols if j > s_idx]
        if not (left_oi and right_oi):
            continue

        price_cols = (_column_positions(cells, COLUMN_HINTS["settle"])
                      or _column_positions(cells, COLUMN_HINTS["close"]))
        vol_cols = _column_positions(cells, COLUMN_HINTS["volume"])
        month = sniff_month(rows[:i + 1]) or default_month
        if not month:
            return None

        chain = sa.Chain(contract_month=month)
        for row2 in rows[i + 1:]:
            if s_idx >= len(row2):
                continue
            strike = _num(row2[s_idx])
            if not strike or strike <= 0:
                continue

            def pick(cols: Sequence[int], left: bool) -> Optional[float]:
                for j in cols:
                    if (j < s_idx) == left and j < len(row2):
                        return _num(row2[j])
                return None

            r = sa.StrikeRow(strike=strike)
            r.call_oi = int(pick(left_oi, True) or 0)
            r.put_oi = int(pick(right_oi, False) or 0)
            r.call_price = pick(price_cols, True)
            r.put_price = pick(price_cols, False)
            r.call_volume = int(pick(vol_cols, True) or 0)
            r.put_volume = int(pick(vol_cols, False) or 0)
            if r.call_oi or r.put_oi or r.call_price or r.put_price:
                chain.add(r)
        return chain if chain.rows else None
    return None


def _chains_from_rows(rows: Sequence[Sequence[str]], raw: bytes, name: str,
                      months: Optional[Sequence[str]],
                      default_month: Optional[str]) -> Dict[str, sa.Chain]:
    header = find_header(rows)
    if header and "kind" in header[1]:
        return parse_rows_long(rows, months)
    wide = parse_wide_chain(rows, default_month)
    if wide:
        return {wide.contract_month: wide}
    return parse_rows_long(rows, months)


def parse_chains_any(raw: bytes, name: str = "",
                     months: Optional[Sequence[str]] = None,
                     default_month: Optional[str] = None) -> Dict[str, sa.Chain]:
    """
    長形式・横形式の両方を試す。
    Excelは限月ごとにシートが分かれていることがあるので、シート単位で解析して束ねる。
    """
    lower = name.lower()
    is_excel = lower.endswith((".xlsx", ".xlsm", ".xls")) or (
        raw[:2] == b"PK" and not lower.endswith(".zip"))
    if is_excel:
        chains: Dict[str, sa.Chain] = {}
        errors: List[str] = []
        for sheet_name, rows in sheets_from_excel(raw):
            try:
                got = _chains_from_rows(rows, raw, sheet_name, months,
                                        sniff_month([[sheet_name]]) or default_month)
            except FetchError as exc:
                errors.append(f"{sheet_name}: {exc}")
                continue
            for m, ch in got.items():
                if m in chains:
                    chains[m].rows.update(ch.rows)
                else:
                    chains[m] = ch
        if chains:
            return chains
        raise FetchError("どのシートからもオプション板を抽出できませんでした。"
                         + ("／".join(errors) if errors else ""))

    return _chains_from_rows(rows_from_bytes(raw, name), raw, name, months, default_month)


# ---------------------------------------------------------------------------
# まとめ: 板データの入手
# ---------------------------------------------------------------------------

@dataclass
class ChainSource:
    chains: Dict[str, sa.Chain]
    origin: str
    as_of: Optional[date] = None
    notes: List[str] = field(default_factory=list)
    meta: Dict[str, "MonthMeta"] = field(default_factory=dict)


def _merge_settlement_prices(chains: Dict[str, sa.Chain],
                             priced: Dict[str, sa.Chain]) -> int:
    """日報PDFで清算価格が取れなかった行を、清算値段CSVの値で埋める。"""
    filled = 0
    for month, ch in chains.items():
        src = priced.get(month)
        if not src:
            continue
        for strike, row in ch.rows.items():
            other = src.rows.get(strike)
            if not other:
                continue
            if row.call_price is None and other.call_price is not None:
                row.call_price = other.call_price
                filled += 1
            if row.put_price is None and other.put_price is not None:
                row.put_price = other.put_price
                filled += 1
    return filled


def load_chains(months: Optional[Sequence[str]] = None,
                target: Optional[date] = None) -> ChainSource:
    """
    オプション板を入手する。優先順位:
      1. SQ_CHAIN_FILE（手元のファイル。カンマ区切りで複数指定可）
      2. 大阪取引所日報 siop_dyr PDF（行使価格別の建玉と清算価格）
         ＋ 同じ営業日の清算値段CSV（原資産価格・金利・残日数の補完）

    行使価格別の建玉は日次では日報PDFにしか無いため、日報を主ソースとする。
    日報は当日中に公開されないことがあるので、公開済みの最新営業日を採用する。
    """
    local = os.environ.get("SQ_CHAIN_FILE")
    if local:
        chains: Dict[str, sa.Chain] = {}
        names: List[str] = []
        for path in [p.strip() for p in local.split(",") if p.strip()]:
            if not os.path.exists(path):
                raise FetchError(f"指定されたファイルが見つかりません: {path}")
            with open(path, "rb") as f:
                raw = f.read()
            got = parse_chains_any(raw, os.path.basename(path), months)
            for m, ch in got.items():
                if m in chains:
                    chains[m].rows.update(ch.rows)
                else:
                    chains[m] = ch
            names.append(os.path.basename(path))
        return ChainSource(chains=chains, origin="手元ファイル: " + ", ".join(names),
                           as_of=target)

    d, chains, pdf_name = fetch_daily_report_chains(target, months)
    notes = [f"建玉・清算価格: 大阪取引所日報 {pdf_name}"]

    # 同じ営業日の清算値段CSVで、原資産価格・金利・残日数を補い、価格の欠けも埋める
    meta: Dict[str, MonthMeta] = {}
    try:
        page = http_get(JPX_SETTLEMENT_PAGE).decode("utf-8", errors="replace")
        url = next((u for dd, u in discover_settlement_files(page) if dd == d), None)
        if url:
            priced, meta = parse_settlement_csv(http_get(url), months)
            filled = _merge_settlement_prices(chains, priced)
            notes.append(f"原資産価格・金利: 清算値段 {url.rsplit('/', 1)[-1]}"
                         + (f"（価格の欠け {filled} 件を補完）" if filled else ""))
        else:
            notes.append(f"⚠ {d} の清算値段CSVが見つからず、原資産価格・金利は未取得")
    except FetchError as exc:
        notes.append(f"⚠ 清算値段CSVを取得できず補完なし: {exc}")

    return ChainSource(chains=chains, origin=f"JPX 大阪取引所日報 {d.isoformat()}",
                       as_of=d, notes=notes, meta=meta)


# ---------------------------------------------------------------------------
# JPX 清算値段CSV（rbYYYYMMDD.csv）
# ---------------------------------------------------------------------------
#
# 実際の中身（2026-08-19 時点で確認）:
#   銘柄コード,銘柄名称,PUT/CAL,限月,権利行使価格,清算価格,理論価格,原資産価格,
#   ボラティリティ,金利,残日数,原資産名称
#   141330018,CAL_225_260910_20000,CAL,202609,20000,45450,45447,65326.42,228.53,1.1529,23,日経225
#
# 銘柄名称は <CAL|PUT>_225_<最終売買日 YYMMDD>_<行使価格>。
# ウィークリーやミニは最終売買日が異なるため、この日付で系列を選り分けられる。
# ボラティリティ列は深いITMで極端な値を取るので採用せず、清算価格から自前で逆算する。

SETTLEMENT_LINK_RE = re.compile(r'href="([^"]*?/(?:rb)(\d{8})\.csv)"', re.IGNORECASE)
NIKKEI_SERIES_RE = re.compile(r"^(CAL|PUT)_225_(\d{6})_(\d+)$")
SETTLE_HEADER_KEY = "銘柄コード"
NIKKEI_UNDERLYING = "日経225"


@dataclass
class MonthMeta:
    """限月ごとに JPX が付けている参考値。"""
    underlying: Optional[float] = None   # 原資産価格
    rate_pct: Optional[float] = None     # 金利（%）
    days: Optional[int] = None           # 残日数
    last_trading_day: Optional[str] = None


def discover_settlement_files(page_html: str) -> List[Tuple[date, str]]:
    """清算値段ページから (日付, 絶対URL) を新しい順に返す。"""
    found: Dict[date, str] = {}
    for href, ymd in SETTLEMENT_LINK_RE.findall(page_html):
        try:
            d = datetime.strptime(ymd, "%Y%m%d").date()
        except ValueError:
            continue
        url = href if href.startswith("http") else JPX_BASE + (
            href if href.startswith("/") else "/" + href)
        found.setdefault(d, url)
    return sorted(found.items(), key=lambda x: x[0], reverse=True)


def parse_settlement_csv(raw: bytes, months: Optional[Sequence[str]] = None
                         ) -> Tuple[Dict[str, sa.Chain], Dict[str, MonthMeta]]:
    """清算値段CSVから日経225オプション（標準限月）の板を組み立てる。"""
    rows = rows_from_bytes(raw, "settlement.csv")
    hidx = next((i for i, r in enumerate(rows)
                 if r and str(r[0]).strip() == SETTLE_HEADER_KEY), None)
    if hidx is None:
        raise FetchError("清算値段CSVのヘッダ行（銘柄コード…）が見つかりません。")
    header = [str(c).strip() for c in rows[hidx]]
    col = {name: i for i, name in enumerate(header)}
    required = ("銘柄名称", "PUT/CAL", "限月", "権利行使価格", "清算価格", "原資産名称")
    missing = [c for c in required if c not in col]
    if missing:
        raise FetchError(f"清算値段CSVに想定した列がありません: {missing} / 実際: {header}")

    chains: Dict[str, sa.Chain] = {}
    meta: Dict[str, MonthMeta] = {}
    seen_series = 0

    for row in rows[hidx + 1:]:
        def cell(name: str) -> str:
            i = col.get(name)
            return str(row[i]).strip() if i is not None and i < len(row) else ""

        if cell("原資産名称") != NIKKEI_UNDERLYING:
            continue
        m = NIKKEI_SERIES_RE.match(cell("銘柄名称"))
        if not m:
            continue          # ウィークリー・ミニ・その他の系列
        seen_series += 1

        month = _normalise_month(cell("限月"))
        if month is None or (months and month not in months):
            continue
        strike = _num(cell("権利行使価格"))
        price = _num(cell("清算価格"))
        if not strike or strike <= 0:
            continue
        is_call = m.group(1) == "CAL"

        chain = chains.setdefault(month, sa.Chain(contract_month=month))
        r = chain.rows.get(strike) or sa.StrikeRow(strike=strike)
        if is_call:
            r.call_price = price
        else:
            r.put_price = price
        chain.add(r)

        mm = meta.setdefault(month, MonthMeta())
        if mm.underlying is None:
            mm.underlying = _num(cell("原資産価格"))
            mm.rate_pct = _num(cell("金利"))
            d = _num(cell("残日数"))
            mm.days = int(d) if d is not None else None
            mm.last_trading_day = m.group(2)

    if not chains:
        raise FetchError(
            "清算値段CSVから日経225オプションの標準限月を抽出できませんでした"
            f"（該当形式の銘柄 {seen_series} 件）。")
    return chains, meta


# ---------------------------------------------------------------------------
# 大阪取引所日報（株価指数オプション）siop_dyr_YYYYMMDD.pdf
# ---------------------------------------------------------------------------
#
# 行使価格別の建玉残高は、日次ではこのPDFにしか無い（2026-08 時点で確認）。
# 1行の形は次のとおりで、右端が建玉残高:
#
#   202609 09.10 31,000 131091018 … … … … 3.0000 4.0000 3.0000 4.0000 + 2.0000 23 76,000 4.00 … 242
#   限月   最終日 行使価格 コード  ←夜間4本値→  ←日中4本値→        前日比  出来高 代金  清算値 権利行使 建玉
#
# 前日比は符号が別トークンになったり「…」になったりして列数が動く。
# そこで中間列は解釈せず、左から4つ・右から3つだけを読む。
#
# CALL/PUT はページ内の「プットオプション」「コールオプション」の見出しで切り替わり、
# 見出しはページをまたいで効き続けるので、読み順に状態を持ち越す。

DAILY_REPORT_JSON = "/automation/markets/statistics-derivatives/daily/json/daily_report_{month}.json"
DAILY_REPORT_ZIP = ("/automation/markets/statistics-derivatives/daily/files/"
                    "{month}/Daily_Report_OSE_{date}.zip")

SIOP_LINE_RE = re.compile(
    r"^(\d{6})\s+(\d{1,2}\.\d{2})\s+([\d,]+)\s+(\d{6,12})\s+(.+)$")
MISSING_TOKENS = {"…", "...", "-", "―", "‐"}

CALL_MARKERS = ("コールオプション", "CallOptions", "Call Options")
PUT_MARKERS = ("プットオプション", "PutOptions", "Put Options")

# 商品の判定は「行頭が商品名のタイトル行」で行う。
# ページ冒頭の注記
#   ※日経225オプション、日経225ミニオプションの場合、表示単位「Ｐ」は「円」に置き換え
# には両方の商品名が出てくるので、単なる部分一致で判定すると本体ページごと落ちる。
# タイトル行は各ページに繰り返されるため、ページごとに状態を初期化して拾い直す。
NIKKEI225_OPTION_TITLE = "日経225オプション"


def _siop_number(token: str) -> Optional[float]:
    if token in MISSING_TOKENS:
        return None
    return _num(token)


def parse_siop_pages(pages: Iterable[str],
                     months: Optional[Sequence[str]] = None) -> Dict[str, sa.Chain]:
    """
    日報（株価指数オプション）のページテキストから、日経225オプションの
    行使価格別 建玉残高・清算価格を取り出す。PDFに依存しないのでテストしやすい。
    """
    chains: Dict[str, sa.Chain] = {}
    is_call: Optional[bool] = None   # PUT/CALL の見出しはページをまたいで効き続ける

    for text in pages:
        if not text:
            continue
        # 商品はページごとに判定し直す。タイトル行が無いページは対象外として扱う。
        in_nikkei225 = any(line.strip().startswith(NIKKEI225_OPTION_TITLE)
                           for line in text.splitlines())

        for raw_line in text.splitlines():
            line = raw_line.strip()
            if any(m in line for m in CALL_MARKERS):
                is_call = True
            elif any(m in line for m in PUT_MARKERS):
                is_call = False
            if not in_nikkei225 or is_call is None:
                continue

            m = SIOP_LINE_RE.match(line)
            if not m:
                continue
            month = _normalise_month(m.group(1))
            strike = _num(m.group(3))
            if month is None or not strike or strike <= 0:
                continue
            if months and month not in months:
                continue

            tail = m.group(5).split()
            if len(tail) < 3:
                continue
            oi = _siop_number(tail[-1])
            settle = _siop_number(tail[-3])
            if oi is None and settle is None:
                continue

            chain = chains.setdefault(month, sa.Chain(contract_month=month))
            row = chain.rows.get(strike) or sa.StrikeRow(strike=strike)
            if is_call:
                row.call_oi = int(oi or 0)
                if settle is not None:
                    row.call_price = settle
            else:
                row.put_oi = int(oi or 0)
                if settle is not None:
                    row.put_price = settle
            chain.add(row)

    return chains


def parse_siop_pdf(pdf_bytes: bytes,
                   months: Optional[Sequence[str]] = None) -> Dict[str, sa.Chain]:
    try:
        import pdfplumber
    except ImportError as exc:
        raise FetchError("日報PDFを読むには pdfplumber が必要です。") from exc
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        pages = [p.extract_text() or "" for p in pdf.pages]
    chains = parse_siop_pages(pages, months)
    if not chains:
        raise FetchError(
            f"日報PDF（{len(pages)}ページ）から日経225オプションの建玉を抽出できませんでした。")
    return chains


def daily_report_url(trade_date: date) -> str:
    return JPX_BASE + DAILY_REPORT_ZIP.format(
        month=trade_date.strftime("%Y%m"), date=trade_date.strftime("%Y%m%d"))


def latest_daily_report(target: Optional[date] = None,
                        lookback_months: int = 2) -> Tuple[date, str]:
    """
    日報が公開されている最新の営業日を、月次JSONから調べる。
    日報は当日中に出ないことがあるため、公開済みの最新日を採用する。
    """
    import json as _json
    months: List[str] = []
    base = target or date.today()
    for i in range(lookback_months):
        y, m = base.year, base.month - i
        while m <= 0:
            y, m = y - 1, m + 12
        months.append(f"{y}{m:02d}")

    for month in months:
        try:
            raw = http_get(JPX_BASE + DAILY_REPORT_JSON.format(month=month), retries=2)
            data = _json.loads(raw.decode("utf-8", errors="replace"))
        except (FetchError, ValueError):
            continue
        entries = []
        for row in data.get("TableDatas", []):
            td = str(row.get("TradeDate") or "")
            if len(td) != 8 or not td.isdigit():
                continue
            try:
                d = datetime.strptime(td, "%Y%m%d").date()
            except ValueError:
                continue
            path = row.get("OseAll")
            if path:
                entries.append((d, JPX_BASE + path if path.startswith("/") else path))
        entries.sort(key=lambda x: x[0], reverse=True)
        if target:
            for d, url in entries:
                if d == target:
                    return d, url
        elif entries:
            return entries[0]
    raise FetchError("日報（Daily_Report_OSE）の公開一覧を取得できませんでした。")


def fetch_daily_report_chains(target: Optional[date] = None,
                              months: Optional[Sequence[str]] = None
                              ) -> Tuple[date, Dict[str, sa.Chain], str]:
    """日報ZIPを取得し、株価指数オプションのPDFから板を組み立てる。"""
    d, url = latest_daily_report(target)
    raw = http_get(url)
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        name = next((n for n in zf.namelist()
                     if n.startswith("siop_dyr_") and "flex" not in n), None)
        if not name:
            raise FetchError(f"日報ZIPに株価指数オプションのPDFがありません: {zf.namelist()}")
        pdf_bytes = zf.read(name)
    return d, parse_siop_pdf(pdf_bytes, months), name
