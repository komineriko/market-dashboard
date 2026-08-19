#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""行使価格別の建玉が取れるかを最終確認する。日報ZIPの中身と、建玉残高xlsxの構造を見る。"""
import io
import sys
import zipfile

import requests

UA = "Mozilla/5.0 (compatible; market-dashboard/1.0; +https://github.com/komineriko/market-dashboard)"
BASE = "https://www.jpx.co.jp"

# 直近の営業日をいくつか試す（当日分がまだ無い場合に備える）
DATES = ["20260819", "20260818", "20260817"]


def get(url):
    return requests.get(url, headers={"User-Agent": UA}, timeout=120)


def decode(raw):
    for enc in ("utf-8", "cp932"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("cp932", errors="replace")


def check_daily_report_zip():
    print(f"\n{'=' * 78}\n### 大阪取引所日報 ZIP の中身")
    for d in DATES:
        url = f"{BASE}/automation/markets/statistics-derivatives/daily/files/{d[:6]}/Daily_Report_OSE_{d}.zip"
        try:
            r = get(url)
        except Exception as exc:                   # noqa: BLE001
            print(f"  {d}: 取得失敗 {exc}")
            continue
        print(f"  {d}: HTTP {r.status_code} / {len(r.content):,} bytes")
        if r.status_code != 200:
            continue
        with zipfile.ZipFile(io.BytesIO(r.content)) as zf:
            for info in zf.infolist():
                print(f"      {info.filename}  ({info.file_size:,} bytes)")
            # PDF以外があれば中身を見る
            for info in zf.infolist():
                if not info.filename.lower().endswith(".pdf"):
                    raw = zf.read(info.filename)
                    print(f"      --- {info.filename} の先頭 ---")
                    for line in decode(raw).splitlines()[:12]:
                        print(f"      | {line[:260]}")
        return


def check_oi_xlsx():
    print(f"\n{'=' * 78}\n### 日経225オプション 建玉残高 xlsx の構造")
    from openpyxl import load_workbook
    for d in DATES:
        url = f"{BASE}/automation/markets/derivatives/open-interest/files/{d[:4]}/{d}_nk225op_oi_by_tp.xlsx"
        try:
            r = get(url)
        except Exception as exc:                   # noqa: BLE001
            print(f"  {d}: 取得失敗 {exc}")
            continue
        print(f"  {d}: HTTP {r.status_code} / {len(r.content):,} bytes  ({url})")
        if r.status_code != 200:
            continue
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        print(f"      シート: {wb.sheetnames}")
        for ws in wb.worksheets[:3]:
            print(f"      --- シート「{ws.title}」 ({ws.max_row} 行 x {ws.max_column} 列) ---")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                cells = ["" if c is None else str(c) for c in row]
                print(f"      | {' | '.join(cells)[:280]}")
                if i >= 14:
                    break
        return


def main() -> int:
    check_daily_report_zip()
    try:
        check_oi_xlsx()
    except Exception as exc:                       # noqa: BLE001
        print(f"  建玉xlsxの確認で例外: {exc}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
